from PIL import Image as Img
from wand.image import Image
import glob
import pyocr
import pyocr.builders
import os
import sys
import codecs
import openpyxl
import pprint



def pdf2png():
    pNum = 0
    for pdfs in glob.glob("*.pdf"):
        #quality is very important (we can test the OCR with lower qualities to increase speed)
        with Image(filename=pdfs,resolution=300) as img:
            pNum +=1
            img.compression_quality = 80
            
            #we have to take out the .pdf out of all of these
            image_name = pdfs[0:len(pdfs)-4]
            
            #straight up changing the name
            img.save(filename=image_name + str(pNum)+ ".png")
            print(image_name+ " " + str(pNum)+ " was changed from .pdf to .png")

#pdf2png()

def image2text():
    
    #all pngs in a list
    image_file = glob.glob("*.png")
    
    #identifying page numbers
    identity = 0
    
    for image in image_file:
        identity += 1
        
        im = Img.open(image)
        tool = pyocr.get_available_tools()[0]
        lang = tool.get_available_languages()[1]
        text_r = tool.image_to_string((im),lang='eng',builder=pyocr.builders.TextBuilder())
        
        file = codecs.open(image[0:len(image)-4] + str(identity) + ".txt", "w", "utf-8")
        file.write(text_r)
        file.close()
        os.unlink(image)

#image2text()

def search_for():
    for filename in glob.glob("*.txt"):
        count = 0
        with codecs.open(filename, 'r', "utf-8") as inF:
            for line in inF:
                #print line
                if '[nc]' in line.lower():
                    count +=1
                if '(nc)' in line.lower():
                    count +=1
                if '{nc}' in line.lower():
                    count +=1
                if 'newscanada' in line.lower():
                    count +=1
        printed_name = filename[:-4]
        print("Found " + str(count) + " in " + printed_name)

def readfromExcel():
    #Input is hardcoded rn
    wb = openpyxl.load_workbook('keywords.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')

    values = []

    for i in range(1, sheet.max_row):
        values.append(sheet.cell(row = i, column = 1).value)


    print('Writing results...')
    resultFile = open('keywords.txt', 'w')
    for eachVal in values:
        #print type(eachVal)
        resultFile.write(eachVal.encode('utf-8') + "\n")
    
    resultFile.close()
    print('Done.')

readfromExcel()

def pullKeywordsSearch():
    with codecs.open("keywords.txt", 'r', "utf-8") as inT:
        for eachquery in inT:
            length = len(eachquery)
            wordsinQuery = []
            match = length // 5
            charCount = 0
            startWord = -1
            #make it lowercase and take away spaces
            #print str(eachquery)
            eachquery.lower()
            #print (eachquery)
            for char in eachquery:
                
                #print(charCount)
                if char == " ":
  
                    
                    if eachquery[startWord+1:charCount] == "and" or eachquery[startWord+1:charCount] == "of" or eachquery[startWord+1:charCount] == "is" or eachquery[startWord+1:charCount] == "it" or eachquery[startWord+1:charCount] == "we" or eachquery[startWord+1:charCount] == "i" or eachquery[startWord+1:charCount] == "a" or eachquery[startWord+1:charCount] == "to" or eachquery[startWord+1:charCount] == "the" or eachquery[startWord+1:charCount] == "on" or eachquery[startWord+1:charCount] == "for" or eachquery[startWord+1:charCount] == "in" or eachquery[startWord+1:charCount] == "as" or eachquery[startWord+1:charCount] == "at" or eachquery[startWord+1:charCount] == "by" or eachquery[startWord+1:charCount] == "are" or eachquery[startWord+1:charCount] == "or" or eachquery[startWord+1:charCount] == "all" or eachquery[startWord+1:charCount] == "do":
                        startWord = charCount
                    else:
                        wordsinQuery.append(eachquery[startWord+1:charCount])
                        #print(eachquery[startWord+1:charCount])
                        startWord = charCount
                    

                charCount += 1
            numFound = 0
            numOfWords = len(wordsinQuery)
            for filename in glob.glob("*.txt"):
                #print(filename)
                
                if filename == "keywords.txt":
                    pass
                else:
                    with codecs.open(filename, 'r', "utf-8") as inF:
                        matches = 0
                        matchWords = []
                        fullLine = ""
                        
                        for line in inF:

                            line.lower()
                            line.replace(" ", "")
                            fullLine = fullLine + line
                                
                            

                            
                        for query in wordsinQuery:
                            if query == "and" or query == "of" or query == "is" or query == "it" or query == "we" or query == "i" or query == "a" or query == "to" or query == "the" or query == "on" or query == "for" or query == "in" or query == "as" or query == "at" or query == "by" or query == "are" or query == "or" or query == "all" or query == "do":
                                pass
                            else:
                                location = ((fullLine.lower()).replace(" ","")).find(query)
                                if location != -1:

                                    matchWords.append(query)


                                    matches += 1
                                    #print(numOfWords//3)
                                    if matches == (numOfWords//4):
                                        
                                        #print("Found something:")
                                        
                                        #print(location)
                                        
                                        #print(matchWords)
                                        #print(eachquery)
                                        #print("The file says:")
                                        #print(fullLine)
                                        numFound += 1
            print ("WE FOUND THIS MANY!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!" + str(numFound) + " IN " + filename)

pullKeywordsSearch()


"""
def cleanText():
    import glob

    for files in glob.glob('*.txt'):
        with open(files, 'r') as checks:
            fixedStrings = ""
            locationDash = 0
            locationQuote = 0
            locationWeirdDash = 0
            for lines in checks:
                locationDash = lines.find('-')
                if lines[locationDash + 1].isdigit() == True:
                    #print("ITS TRUE MAKE SHIT HAPPEN DIFFERENT")
                    fixedChars = ""
                    for character in lines:
                        if character.find('-') != -1 and lines[lines.index(character)+1].isdigit() == True:
                            #print("THIS IS THE ONE WE WANT TO KEEP")
                            fixedChars = fixedChars + character
                        else:
                            fixedChars = fixedChars + character.replace("-", "")

                    fixedStrings = fixedStrings + fixedChars
                
                else:
                    fixedStrings = fixedStrings + lines.replace("-", "")

                print (lines)
                
            print(fixedStrings)
cleanText()
                    
    
            
    
    #search from txt
"""
